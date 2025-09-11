import pandas as pd
import numpy as np
import standard
import streamlit as st
from io import BytesIO
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

def getFileDetails(files):
  file_details = []
  for file in files:
    [name, date] = file.name.split("-")
    file_details.append({
      "Customer Name": name,
      "Date": date.replace(".xlsx", "").replace(".xls", ""),
    })
  file_details = pd.DataFrame(file_details)
  return file_details

def getSumGivenColumn(Pline, dataframe, column):
  if Pline == 'All':
    return dataframe[column].sum()
  else:
    if 'Pline' in dataframe.columns:
      filtered_data = dataframe[dataframe['Pline'] == Pline]
    elif 'P Line' in dataframe.columns:
      filtered_data = dataframe[dataframe['P Line'] == Pline]
    else:
      raise KeyError("Neither 'Pline' nor 'P Line' column found in dataframe.")
    total_gross = filtered_data[column].sum()
    return total_gross
  
def getQTYTotalAndDefect(dataframe):
  cumulative_columns = [col for col in dataframe.columns if 'Cumulative' in col]
  for column in cumulative_columns:
    QTY_gross = dataframe.loc[dataframe['Metric'] == 'QTY Gross', column].iloc[0]
    QTY_defect = dataframe.loc[dataframe['Metric'] == 'QTY Defect', column].iloc[0]
    dataframe.loc[dataframe['Metric'] == 'QTY Total', column] = QTY_gross + QTY_defect
    dataframe.loc[dataframe['Metric'] == 'Defect %', column] = QTY_defect / QTY_gross if QTY_gross != 0 else 0
  return dataframe

def getSumAndPerUnit(Pline, dataframe, column):
  if Pline == 'All':
    total = dataframe[column].sum()
    per_unit = total / dataframe['L12 Shipped'].sum() if dataframe['L12 Shipped'].sum() != 0 else 0
    return total, per_unit
  else:
    filtered_data = dataframe[dataframe['Pline'] == Pline] or dataframe[dataframe['P Line'] == Pline]
    total = filtered_data[column].sum()
    per_unit = total / filtered_data['L12 Shipped'].sum() if filtered_data['L12 Shipped'].sum() != 0 else 0
    return total, per_unit

# STANDARD HELPER
def extract_parentheses_content(text):
  start = text.find('(')
  end = text.find(')')
  if start != -1 and end != -1 and start < end:
    content = text[start+1:end]
    return [item.strip() for item in content.split(',')]
  return []

# STANDARD HELPER
def getPerUnit(row, column, output):
  return output.loc[row, column] / output.loc['QTY Gross', column]

# STANDARD
def QTY_CALCULATIONS(PRODUCT_LINES, output, DATA, ASSUMPTIONS, volume):
  total_gross = 0
  for line in PRODUCT_LINES:
    gross = getSumGivenColumn(line, DATA, 'Qty')
    output.loc['QTY Gross', f'{line} Cumulative'] = gross + (gross*volume)
    total_gross += output.loc['QTY Gross', f'{line} Cumulative']
    try:
      output.loc['Defect %', f'{line} Cumulative'] = ASSUMPTIONS.loc[line, 'Defect %']
    except:
      output.loc['Defect %', f'{line} Cumulative'] = ASSUMPTIONS.loc['-', 'Defect %']

  output.loc['QTY Gross', 'All Lines Cumulative'] = total_gross
  total_defect = 0
  total_total = 0
  for line in PRODUCT_LINES:
    output.loc['QTY Defect', f'{line} Cumulative'] = output.loc['QTY Gross', f'{line} Cumulative'] * output.loc['Defect %', f'{line} Cumulative'] * -1
    output.loc['QTY Total', f'{line} Cumulative'] = output.loc['QTY Gross', f'{line} Cumulative'] + output.loc['QTY Defect', f'{line} Cumulative']
    total_defect += output.loc['QTY Defect', f'{line} Cumulative']
    total_total += output.loc['QTY Total', f'{line} Cumulative']
  output.loc['QTY Defect', 'All Lines Cumulative'] = total_defect
  output.loc['QTY Total', 'All Lines Cumulative'] = total_total
  defect_percent = total_defect / total_gross
  output.loc['Defect %', 'All Lines Cumulative'] = defect_percent * -1
  return output

# STANDARD
def GET_NETSALES_MARGIN_METRICS(FORMAT):
  defect_idx = FORMAT.index.get_loc('Defect %')
  net_sales_idx = FORMAT.index.get_loc('NET SALES')
  total_variable_cost_idx = FORMAT.index.get_loc('TOTAL VARIABLE COST')

  net_sales_metrics = FORMAT.index[defect_idx+1:net_sales_idx]
  net_sales_list = net_sales_metrics.tolist()
  if 'Agency Rep' in net_sales_list:
    net_sales_list.remove('Agency Rep')
    net_sales_list.append('Agency Rep')
    net_sales_metrics = pd.Index(net_sales_list)

  margin_metrics = FORMAT.index[net_sales_idx+1:total_variable_cost_idx]
  margin_list = margin_metrics.tolist()
  if 'Scrap Return Rate' in margin_list:
    margin_list.remove('Scrap Return Rate')
    margin_list.append('Scrap Return Rate')
    margin_metrics = pd.Index(margin_list)

  return net_sales_metrics, margin_metrics

# STANDARD
def GET_SGA_METRICS(FORMAT):
  margin_idx = FORMAT.index.get_loc('MARGIN %')
  sga_idx = FORMAT.index.get_loc('SG&A')
  sga_metrics = FORMAT.index[margin_idx+1:sga_idx]
  return sga_metrics.tolist()

def NET_SALES_CALCULATIONS(output, DATA, ASSUMPTIONS, NET_SALES_METRICS, PRODUCT_LINES, DEFAULTS, volume):
  for metric in NET_SALES_METRICS:
    total_metric = 0
    specific_pline = extract_parentheses_content(metric)
    if metric == 'Defect':
      for line in PRODUCT_LINES:
        try:
          output.loc[metric, f'{line} Cumulative'] = output.loc['Sales', f'{line} Cumulative'] * ASSUMPTIONS.loc[line, 'Defect %'] * -1
        except:
          output.loc[metric, f'{line} Cumulative'] = output.loc['Sales', f'{line} Cumulative'] * ASSUMPTIONS.loc['-', 'Defect %'] * -1
        total_metric += output.loc[metric, f'{line} Cumulative']
        output.loc[metric, f'{line} Per Unit'] = getPerUnit(metric, f'{line} Cumulative', output)
    elif metric not in DEFAULTS:
      for line in PRODUCT_LINES:
        temp_sale = getSumGivenColumn(line, DATA, metric)
        output.loc[metric, f'{line} Cumulative'] = temp_sale + (temp_sale*volume)
        total_metric += output.loc[metric, f'{line} Cumulative']
        output.loc[metric, f'{line} Per Unit'] = getPerUnit(metric, f'{line} Cumulative', output)
    else:
      loss = -1
      for line in PRODUCT_LINES:
        if line in specific_pline or specific_pline == []:
          output.loc[metric, f'{line} Cumulative'] = output.loc['Sales', f'{line} Cumulative'] * DEFAULTS.get(metric) * loss
          total_metric += output.loc[metric, f'{line} Cumulative']
        else:
          output.loc[metric, f'{line} Cumulative'] = 0
        output.loc[metric, f'{line} Per Unit'] = getPerUnit(metric, f'{line} Cumulative', output)
    output.loc[metric, 'All Lines Cumulative'] = total_metric
  total_metric = 0
  for line in PRODUCT_LINES:
    agency_rep = 0
    for metric in NET_SALES_METRICS:
      if metric != 'Agency Rep':
        agency_rep += output.loc[metric, f'{line} Cumulative']
    output.loc['Agency Rep', f'{line} Cumulative'] = agency_rep * -1 * DEFAULTS.get('Agency Rep', 0)
    total_metric += output.loc['Agency Rep', f'{line} Cumulative']
    output.loc['Agency Rep', f'{line} Per Unit'] = getPerUnit('Agency Rep', f'{line} Cumulative', output)
  output.loc['Agency Rep', 'All Lines Cumulative'] = total_metric
  total_net_sales = 0
  for line in PRODUCT_LINES:
    net_sales = 0
    for metric in NET_SALES_METRICS:
      net_sales += output.loc[metric, f'{line} Cumulative']
    output.loc['NET SALES', f'{line} Cumulative'] = net_sales
    total_net_sales += net_sales
    output.loc['NET SALES', f'{line} Per Unit'] = getPerUnit('NET SALES', f'{line} Cumulative', output)

  output.loc['NET SALES', 'All Lines Cumulative'] = total_net_sales
  for metric in NET_SALES_METRICS:
    output.loc[metric, 'All Lines Per Unit'] = getPerUnit(metric, 'All Lines Cumulative', output)
  output.loc['NET SALES', 'All Lines Per Unit'] = getPerUnit('NET SALES', 'All Lines Cumulative', output)
  return output

# STANDARD
def MARGIN_CALCULATIONS(PRODUCT_LINES, MARGIN_METRICS, output, DATA, DEFAULTS, FORMAT):
  for metric in MARGIN_METRICS:
    output.loc[metric, 'All Lines Cumulative'] = 0
  for line in PRODUCT_LINES:
    for metric in MARGIN_METRICS:
      try:
        if metric == 'Cost':
          output.loc[metric, f'{line} Cumulative'] = getSumGivenColumn(line, DATA, 'Total Cost')
          output.loc[metric, 'All Lines Cumulative'] += output.loc[metric, f'{line} Cumulative']
        else:
          output.loc[metric, f'{line} Cumulative'] = getSumGivenColumn(line, DATA, metric)
          output.loc[metric, 'All Lines Cumulative'] += output.loc[metric, f'{line} Cumulative']
      except:
        output.loc[metric, f'{line} Cumulative'] = 0
      output.loc[metric, f'{line} Per Unit'] = getPerUnit(metric, f'{line} Cumulative', output)
      output.loc[metric, 'All Lines Per Unit'] = getPerUnit(metric, 'All Lines Cumulative', output)

  total_metric = 0
  for line in PRODUCT_LINES:
    scrap_return = 0
    for metric in MARGIN_METRICS:
      if metric != 'Scrap Return Rate':
        if FORMAT.loc[metric].iloc[1] == 1:
          scrap_return += output.loc[metric, f'{line} Per Unit']
    return_allowance = output.loc['Return Allowance', f'{line} Cumulative']
    per_unit_sales = output.loc['NET SALES', f'{line} Per Unit']
    output.loc['Scrap Return Rate', f'{line} Cumulative'] = (1-DEFAULTS.get('Scrap Return Rate'))*(return_allowance/per_unit_sales)*scrap_return
    total_metric += output.loc['Scrap Return Rate', f'{line} Cumulative']
    output.loc['Scrap Return Rate', f'{line} Per Unit'] = getPerUnit('Scrap Return Rate', f'{line} Cumulative', output)
  output.loc['Scrap Return Rate', 'All Lines Cumulative'] = total_metric
  output.loc['Scrap Return Rate', 'All Lines Per Unit'] = getPerUnit('Scrap Return Rate', 'All Lines Cumulative', output)

  total_variable_cost = 0
  total_margin = 0
  for line in PRODUCT_LINES:
    variable_cost = 0
    for metric in MARGIN_METRICS:
      variable_cost += output.loc[metric, f'{line} Cumulative']
    output.loc['TOTAL VARIABLE COST', f'{line} Cumulative'] = variable_cost
    total_variable_cost += variable_cost
    output.loc['TOTAL VARIABLE COST', f'{line} Per Unit'] = getPerUnit('TOTAL VARIABLE COST', f'{line} Cumulative', output)
    net_sales = output.loc['NET SALES', f'{line} Cumulative']
    margin = net_sales - output.loc['TOTAL VARIABLE COST', f'{line} Cumulative']
    output.loc['MARGIN', f'{line} Cumulative'] = margin
    total_margin += margin
    output.loc['MARGIN', f'{line} Per Unit'] = getPerUnit('MARGIN', f'{line} Cumulative', output)
    margin_perc = margin / net_sales
    output.loc['MARGIN %', f'{line} Cumulative'] = margin_perc

  output.loc['TOTAL VARIABLE COST', 'All Lines Cumulative'] = total_variable_cost
  output.loc['TOTAL VARIABLE COST', 'All Lines Per Unit'] = getPerUnit('TOTAL VARIABLE COST', 'All Lines Cumulative', output)
  output.loc['MARGIN', 'All Lines Cumulative'] = total_margin
  output.loc['MARGIN', 'All Lines Per Unit'] = getPerUnit('MARGIN', 'All Lines Cumulative', output)
  output.loc['MARGIN %', 'All Lines Cumulative'] = output.loc['MARGIN', 'All Lines Cumulative'] / output.loc['NET SALES', 'All Lines Cumulative']

  return output

# STANDARD
def SGA_CALCULATIONS(PRODUCT_LINES, output, ASSUMPTIONS, DEFAULTS, SGA_METRICS, file):

  for metric in SGA_METRICS:
    output.loc[metric, 'All Lines Cumulative'] = 0
  output.loc['SG&A', 'All Lines Cumulative'] = 0
  for line in PRODUCT_LINES:
    for metric in SGA_METRICS:
      if metric == 'Fill Rate Fines':
        output.loc['Fill Rate Fines', f'{line} Cumulative'] = output.loc['NET SALES', f'{line} Cumulative'] * DEFAULTS.get('Fill Rate Fines', 0)
        output.loc['Fill Rate Fines', 'All Lines Cumulative'] += output.loc['Fill Rate Fines', f'{line} Cumulative']
        output.loc['Fill Rate Fines', f'{line} Per Unit'] = getPerUnit('Fill Rate Fines', f'{line} Cumulative', output)
      elif metric == 'Inspect Return':
        output.loc['Inspect Return', f'{line} Per Unit'] = output.loc['Handling/Shipping', f'{line} Per Unit']
      elif metric == 'Return Allowance Put Away/Rebox':
        output.loc['Return Allowance Put Away/Rebox', f'{line} Per Unit'] = output.loc['Handling/Shipping', f'{line} Per Unit'] + ASSUMPTIONS.loc[line, 'Return Allowance Put Away/Rebox']
      else:
        output.loc[metric, f'{line} Per Unit'] = ASSUMPTIONS.loc[line, metric]
        output.loc[metric, f'{line} Cumulative'] = ASSUMPTIONS.loc[line, metric] * output.loc['QTY Gross', f'{line} Cumulative']
        output.loc[metric, 'All Lines Cumulative'] += output.loc[metric, f'{line} Cumulative']

    return_allowance = output.loc['Return Allowance', f'{line} Cumulative']
    per_unit_sales = output.loc['NET SALES', f'{line} Per Unit']
    per_unit_inspect = output.loc['Inspect Return', f'{line} Per Unit']
    per_unit_rebox = output.loc['Return Allowance Put Away/Rebox', f'{line} Per Unit']
    if DEFAULTS.get('Scrap Return Rate') == 1:
      output.loc['Inspect Return', f'{line} Cumulative'] = 0
    else:
      output.loc['Inspect Return', f'{line} Cumulative'] = (return_allowance/per_unit_sales)*per_unit_inspect*-1
      output.loc['Inspect Return', 'All Lines Cumulative'] += output.loc['Inspect Return', f'{line} Cumulative']

    output.loc['Return Allowance Put Away/Rebox', f'{line} Cumulative'] = (return_allowance/per_unit_sales)*per_unit_rebox*-1*(1-DEFAULTS.get('Scrap Return Rate', 0))
    output.loc['Return Allowance Put Away/Rebox', 'All Lines Cumulative'] += output.loc['Return Allowance Put Away/Rebox', f'{line} Cumulative']
    
    SGA = 0
    for metric in SGA_METRICS:
      SGA += output.loc[metric, f'{line} Cumulative']
    output.loc['SG&A', f'{line} Cumulative'] = SGA
    output.loc['SG&A', 'All Lines Cumulative'] += SGA
    output.loc['SG&A', f'{line} Per Unit'] = getPerUnit('SG&A', f'{line} Cumulative', output)

  for metric in SGA_METRICS:
    output.loc[metric, 'All Lines Per Unit'] = getPerUnit(metric, 'All Lines Cumulative', output)
  output.loc['SG&A', 'All Lines Per Unit'] = getPerUnit('SG&A', 'All Lines Cumulative', output)

  return output

class ExcelExport:
  def table_format(x, metric):
    try:
      num = float(x)
      if np.isnan(num):
        return ""
      if metric == "Defect %" or metric == "MARGIN %" or metric == "Contribution Margin %":
        if num < 0:
          return f"{abs(num*100):,.2f} %"
        return f"{num*100:,.2f} %"
      elif metric == "QTY Gross" or metric == "QTY Defect" or metric == "QTY Total":
        if num < 0:
          return f'{(int(num)):,}'
        return f'{(int(num)):,}'
      if num < 0:
        return f"$ ({abs(num):,.2f})"
      return f"$ {num:,.2f}"
    except:
      return x
    
  def highlight_negative(val):
    try:
      num = float(val.replace('%','').replace('$','').replace(',','').replace('(','-').replace(')',''))
      color = '#FFECEC' if num < 0 else '#E8F9EE'
      if num == 0:
        color = '#FFFCE7'
    except:
      color = '#E8F2FC'
    return f'background-color: {color}'
  
  def overallBCA(summary_df, input_df, summary_defaults=None, input_defaults=None, plines=None, metrics=None):
    output = BytesIO()
    main_delta = input_df.iloc[:, 1:] - summary_df.iloc[:, 1:]
    delta_dataframe = pd.DataFrame(
        main_delta.values,
        index=input_df.iloc[:,0],
        columns=input_df.columns[1:,]
    )
    delta = delta_dataframe.reset_index()
    delta.rename(columns={'index': 'Metric'}, inplace=True)
    if summary_defaults is not None and input_defaults is not None:
      delta_defaults = input_defaults - summary_defaults
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
      sheetNames = ['Original BCA', 'Scenario BCA', 'Delta BCA', 'Original Defaults', 'Scenario Defaults', 'Delta Defaults']
      ExcelExport.FormatBCA(summary_df, summary_defaults, sheetNames[0], sheetNames[3], writer, plines, metrics)
      ExcelExport.FormatBCA(input_df, input_defaults, sheetNames[1], sheetNames[4], writer, plines, metrics)
      ExcelExport.FormatBCA(delta, delta_defaults, sheetNames[2], sheetNames[5], writer, plines, metrics)
      for sheetName in sheetNames:
        worksheet = writer.sheets[sheetName]
        if 'BCA' in sheetName:
          worksheet.freeze_panes = 'C3'
        else:
          worksheet.freeze_panes = 'A2'
    output.seek(0)
    return output

  def addSpaces(df):
    export_df = df.copy()
    
    if 'Metric' in export_df.columns:
      metric_col_idx = export_df.columns.get_loc('Metric')
      export_df.insert(metric_col_idx + 1, ' ', '')
      
      original_data_cols = [col for col in df.columns if col != 'Metric']
      
      blank_col_counter = 1
      for i in range(2, len(original_data_cols), 2):
        target_col = original_data_cols[i]
        current_col_idx = export_df.columns.get_loc(target_col)
        blank_col_name = ' ' * (blank_col_counter + 1)
        export_df.insert(current_col_idx, blank_col_name, '')
        blank_col_counter += 1
    
    blank_row_triggers = ['Defect %', 'NET SALES', 'MARGIN %', 'SG&A', 'FACTORING %']
    rows_to_insert = []
    
    for idx in export_df.index:
      if 'Metric' in export_df.columns:
        metric_value = str(export_df.loc[idx, 'Metric'])
      else:
        metric_value = str(idx)
      
      if any(trigger in metric_value for trigger in blank_row_triggers):
        rows_to_insert.append(idx)
    
    for row_idx in reversed(rows_to_insert):
      blank_row = pd.Series([''] * len(export_df.columns), index=export_df.columns)
      export_df = pd.concat([
        export_df.iloc[:export_df.index.get_loc(row_idx) + 1],
        pd.DataFrame([blank_row]),
        export_df.iloc[export_df.index.get_loc(row_idx) + 1:]
      ], ignore_index=True)
    return export_df
  
  def FormatBCA(dataframe, default_dataframe, bca_sheetname, defaults_sheetname, writer, plines, metrics):
    export_df = ExcelExport.addSpaces(dataframe)
    export_df.to_excel(writer, index=False, sheet_name=bca_sheetname)
    worksheet = writer.sheets[bca_sheetname]
    worksheet.insert_rows(1)
    for row in worksheet.iter_rows():
      for cell in row:
        cell.border = Border()
    data_cols = [col for col in export_df.columns if col != 'Metric' and not col.strip() == '']
    product_lines = {}
    for i, col in enumerate(data_cols):
      if 'Cumulative' in col or 'Per Unit' in col:
        if 'Cumulative' in col:
          product_name = col.replace(' Cumulative', '').strip()
        else:
          product_name = col.replace(' Per Unit', '').strip()
        if product_name not in product_lines:
          product_lines[product_name] = []
        
        col_pos = export_df.columns.get_loc(col) + 1
        product_lines[product_name].append(col_pos)

    full_border = Border(top=Side(style='thin'), bottom=Side(style='thin'), left=Side(style='thin'), right=Side(style='thin'))

    for product_name, col_positions in product_lines.items():
      if len(col_positions) >= 2:
        start_col = min(col_positions)
        end_col = max(col_positions)
        
        start_cell = f"{get_column_letter(start_col)}1"
        end_cell = f"{get_column_letter(end_col)}1"
        worksheet.merge_cells(f"{start_cell}:{end_cell}")
        worksheet[start_cell].alignment = Alignment(horizontal='center', vertical='center')
        worksheet[start_cell] = product_name
        worksheet[start_cell].border = full_border
        worksheet[end_cell].border = full_border
        worksheet[start_cell].font = Font(bold=True, size=12)
    
    for col_idx, col_name in enumerate(export_df.columns, 1):
      cell = worksheet.cell(row=2, column=col_idx)
      if 'Cumulative' in col_name:
        cell.value = 'Cumulative'
        cell.border = full_border
      elif 'Per Unit' in col_name:
        cell.value = 'Per Unit'
        cell.border = full_border
      elif col_name == 'Metric':
        cell.value = 'Metric'
        cell.border = full_border
      else:
        cell.value = col_name
      cell.font = Font(bold=True)
    
    thin_border = Border(
      top=Side(style='thin'),
      bottom=Side(style='thin')
    )
    
    max_row = worksheet.max_row
    max_col = worksheet.max_column
    
    for row in range(1, max_row + 1):
      for col in range(1, max_col + 1):
        if col == 2 and worksheet.cell(row=row, column=1).value in default_dataframe.index:
          worksheet.cell(row=row, column=col).value = default_dataframe.loc[worksheet.cell(row=row, column=1).value].iloc[0]
          worksheet.cell(row=row, column=col).alignment = Alignment(horizontal='center')
        cell = worksheet.cell(row=row, column=col)
        if str(cell.value).strip() != "" and (worksheet.cell(row=row, column=1).value == 'QTY Total'
          or worksheet.cell(row=row, column=1).value == 'NET SALES'
          or worksheet.cell(row=row, column=1).value == 'TOTAL VARIABLE COST'
          or worksheet.cell(row=row, column=1).value == 'MARGIN'
          or worksheet.cell(row=row, column=1).value == 'SG&A'
          or worksheet.cell(row=row, column=1).value == 'Contribution Margin'):
          cell.border = thin_border
        if row <= 2:
          cell.font = Font(bold=True)
        if (worksheet.cell(row=row, column=1).value == 'QTY Total' or 
            worksheet.cell(row=row, column=1).value == 'Defect %' or
            worksheet.cell(row=row, column=1).value == 'NET SALES' or
            worksheet.cell(row=row, column=1).value == 'MARGIN' or 
            worksheet.cell(row=row, column=1).value == 'MARGIN %' or 
            cell.value == 'SG&A' or 
            cell.value == 'FACTORING %' or
            worksheet.cell(row=row, column=1).value == 'Contribution Margin' or
            worksheet.cell(row=row, column=1).value == 'Contribution Margin %'):
          cell.font = Font(bold=True)
        try:
          float(cell.value)
          if (worksheet.cell(row=row, column=1).value == 'Defect %' or
              worksheet.cell(row=row, column=1).value == 'MARGIN %' or
              worksheet.cell(row=row, column=1).value == 'Contribution Margin %' or col == 2):
            cell.number_format = '0.00%'
          elif (worksheet.cell(row=row, column=1).value == 'QTY Gross' or
              worksheet.cell(row=row, column=1).value == 'QTY Defect' or
              worksheet.cell(row=row, column=1).value == 'QTY Total'):
            cell.number_format = '#,##0'
          else:
            cell.number_format = '_($* #,##0.00_);_($* (#,##0.00);_($* "-"??_);_(@_)'
        except:
          pass
    for column in worksheet.columns:
      max_length = 0
      column_letter = get_column_letter(column[0].column)
      for cell in column:
        try:
          if len(str(cell.value)) > max_length:
            max_length = len(str(cell.value))
        except:
          pass
      adjusted_width = min(max_length + 2, 50)
      worksheet.column_dimensions[column_letter].width = adjusted_width
    worksheet.column_dimensions['B'].width = 10

    if default_dataframe is not None:
      new_df = pd.DataFrame(columns=['Product Lines'] + [metric for metric in metrics][::-1])
      new_df['Product Lines'] = [lines for lines in plines]
      for line in plines:
        for metric in metrics:
          new_df.loc[new_df['Product Lines'] == line, metric] = default_dataframe.loc[f'{line} {metric}'].iloc[0]
      new_df.to_excel(writer, sheet_name=defaults_sheetname, index=False)
      worksheet = writer.sheets[defaults_sheetname]
      for row in worksheet.iter_rows():
        for cell in row:
          cell.border = Border()
      for i, col in enumerate(worksheet.iter_cols(min_row=1, max_row=1), 1):
        header = col[0].value
        if header:
          width = min(len(str(header)) + 2, 50)
          worksheet.column_dimensions[get_column_letter(i)].width = width
      max_row = worksheet.max_row
      max_col = worksheet.max_column

      for row in range(1, max_row + 1):
        for col in range(1, max_col + 1):
          worksheet.cell(row=row, column=col).alignment = Alignment(horizontal='center', vertical='center')
          worksheet.cell(row=row, column=col).border = thin_border
          if worksheet.cell(row=1, column=col).value != 'Product Lines':
            if worksheet.cell(row=1, column=col).value != 'Defect %':
              worksheet.cell(row=row, column=col).number_format = '_($* #,##0.00_);_($* (#,##0.00);_($* "-"??_);_(@_)'
            else:
              worksheet.cell(row=row, column=col).number_format = '0.00%'

  def summaryBCA(summary_df, defaults_df=None, plines=None, metrics=None, format_for_export=False):
    output = BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
      export_df = ExcelExport.addSpaces(summary_df)
      export_df.to_excel(writer, index=False, sheet_name='Summary')
      worksheet = writer.sheets['Summary']
      worksheet.insert_rows(1)
      for row in worksheet.iter_rows():
        for cell in row:
          cell.border = Border()
      data_cols = [col for col in export_df.columns if col != 'Metric' and not col.strip() == '']
      product_lines = {}
      for i, col in enumerate(data_cols):
        if 'Cumulative' in col or 'Per Unit' in col:
          if 'Cumulative' in col:
            product_name = col.replace(' Cumulative', '').strip()
          else:
            product_name = col.replace(' Per Unit', '').strip()
          if product_name not in product_lines:
            product_lines[product_name] = []
          
          col_pos = export_df.columns.get_loc(col) + 1
          product_lines[product_name].append(col_pos)

      full_border = Border(top=Side(style='thin'), bottom=Side(style='thin'), left=Side(style='thin'), right=Side(style='thin'))

      for product_name, col_positions in product_lines.items():
        if len(col_positions) >= 2:
          start_col = min(col_positions)
          end_col = max(col_positions)
          
          start_cell = f"{get_column_letter(start_col)}1"
          end_cell = f"{get_column_letter(end_col)}1"
          worksheet.merge_cells(f"{start_cell}:{end_cell}")
          worksheet[start_cell].alignment = Alignment(horizontal='center', vertical='center')
          worksheet[start_cell] = product_name
          worksheet[start_cell].border = full_border
          worksheet[end_cell].border = full_border
          worksheet[start_cell].font = Font(bold=True, size=12)
      
      for col_idx, col_name in enumerate(export_df.columns, 1):
        cell = worksheet.cell(row=2, column=col_idx)
        if 'Cumulative' in col_name:
          cell.value = 'Cumulative'
          cell.border = full_border
        elif 'Per Unit' in col_name:
          cell.value = 'Per Unit'
          cell.border = full_border
        elif col_name == 'Metric':
          cell.value = 'Metric'
          cell.border = full_border
        else:
          cell.value = col_name
        cell.font = Font(bold=True)
      
      thin_border = Border(
        top=Side(style='thin'),
        bottom=Side(style='thin')
      )
      
      max_row = worksheet.max_row
      max_col = worksheet.max_column
      
      for row in range(1, max_row + 1):
        for col in range(1, max_col + 1):
          if col == 2 and worksheet.cell(row=row, column=1).value in defaults_df.index:
            worksheet.cell(row=row, column=col).value = defaults_df.loc[worksheet.cell(row=row, column=1).value].iloc[0]
            worksheet.cell(row=row, column=col).alignment = Alignment(horizontal='center')
          cell = worksheet.cell(row=row, column=col)
          if str(cell.value).strip() != "" and (worksheet.cell(row=row, column=1).value == 'QTY Total'
            or worksheet.cell(row=row, column=1).value == 'NET SALES'
            or worksheet.cell(row=row, column=1).value == 'TOTAL VARIABLE COST'
            or worksheet.cell(row=row, column=1).value == 'MARGIN'
            or worksheet.cell(row=row, column=1).value == 'SG&A'
            or worksheet.cell(row=row, column=1).value == 'Contribution Margin'):
            cell.border = thin_border
          if row <= 2:
            cell.font = Font(bold=True)
          if (worksheet.cell(row=row, column=1).value == 'QTY Total' or 
              worksheet.cell(row=row, column=1).value == 'Defect %' or
              worksheet.cell(row=row, column=1).value == 'NET SALES' or
              worksheet.cell(row=row, column=1).value == 'MARGIN' or 
              worksheet.cell(row=row, column=1).value == 'MARGIN %' or 
              cell.value == 'SG&A' or 
              cell.value == 'FACTORING %' or
              worksheet.cell(row=row, column=1).value == 'Contribution Margin' or
              worksheet.cell(row=row, column=1).value == 'Contribution Margin %'):
            cell.font = Font(bold=True)
          try:
            float(cell.value)
            if (worksheet.cell(row=row, column=1).value == 'Defect %' or
                worksheet.cell(row=row, column=1).value == 'MARGIN %' or
                worksheet.cell(row=row, column=1).value == 'Contribution Margin %' or col == 2):
              cell.number_format = '0.00%'
            elif (worksheet.cell(row=row, column=1).value == 'QTY Gross' or
                worksheet.cell(row=row, column=1).value == 'QTY Defect' or
                worksheet.cell(row=row, column=1).value == 'QTY Total'):
              cell.number_format = '#,##0'
            else:
              cell.number_format = '_($* #,##0.00_);_($* (#,##0.00);_($* "-"??_);_(@_)'
          except:
            pass
      for column in worksheet.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
          try:
            if len(str(cell.value)) > max_length:
              max_length = len(str(cell.value))
          except:
            pass
        adjusted_width = min(max_length + 2, 50)
        worksheet.column_dimensions[column_letter].width = adjusted_width
      worksheet.column_dimensions['B'].width = 10
      
      if defaults_df is not None:
        new_df = pd.DataFrame(columns=['Product Lines'] + [metric for metric in metrics][::-1])
        new_df['Product Lines'] = [lines for lines in plines]
        for line in plines:
          for metric in metrics:
            new_df.loc[new_df['Product Lines'] == line, metric] = defaults_df.loc[f'{line} {metric}'].iloc[0]
        new_df.to_excel(writer, sheet_name='Defaults & Assumptions', index=False)
        worksheet = writer.sheets['Defaults & Assumptions']
        for row in worksheet.iter_rows():
          for cell in row:
            cell.border = Border()
        for i, col in enumerate(worksheet.iter_cols(min_row=1, max_row=1), 1):
          header = col[0].value
          if header:
            width = min(len(str(header)) + 2, 50)
            worksheet.column_dimensions[get_column_letter(i)].width = width
        max_row = worksheet.max_row
        max_col = worksheet.max_column

        for row in range(1, max_row + 1):
          for col in range(1, max_col + 1):
            worksheet.cell(row=row, column=col).alignment = Alignment(horizontal='center', vertical='center')
            worksheet.cell(row=row, column=col).border = thin_border
            if worksheet.cell(row=1, column=col).value != 'Product Lines':
              if worksheet.cell(row=1, column=col).value != 'Defect %':
                worksheet.cell(row=row, column=col).number_format = '_($* #,##0.00_);_($* (#,##0.00);_($* "-"??_);_(@_)'
              else:
                worksheet.cell(row=row, column=col).number_format = '0.00%'
      worksheet.freeze_panes = 'A2'
      worksheet = writer.sheets['Summary']
      worksheet.freeze_panes = 'C3'

    output.seek(0)
    return output.getvalue()
  


class BCA_Matrix:
  def getBCAs(files):
    BCAs = {}
    Plines = []
    Metrics = []
    PerUnitMetrics = []
    customers = st.session_state.fileDetails["Customer Name"].unique()
    for (customer, file) in zip(customers, files):
      info = pd.read_excel(file, sheet_name='Data')
      metrics = pd.read_excel(file, sheet_name='Format')
      Plines.extend(info["P Line"].dropna().unique())
      Metrics.extend(metrics["METRIC"].dropna().unique())
      BCA, _, _, _ = standard.getSummary(file)
      PerUnitMetrics.extend(BCA[~pd.isna(BCA.iloc[:, 2])].loc[:, "Metric"].tolist())
      st.session_state[customer] = BCA.set_index("Metric")
    return BCAs, list(set(Plines)), list(set(Metrics)), list(set(PerUnitMetrics))
  
  def getDescripancy(comparison):
    outlier_result = {}
    customers = list(comparison['Customer'])
    for pline in comparison.columns[1:]:
      values = []
      cust_names = []
      for idx, v in enumerate(comparison[pline]):
        if v != '-' and v != '':
          try:
            val = float(str(v).replace('%','').replace(',','').replace('$','').replace('(','-').replace(')',''))
            values.append(val)
            cust_names.append(customers[idx])
          except:
            continue
      outliers = []
      for idx, val in enumerate(values):
        lower = min(val * 0.95, val * 1.05)
        upper = max(val * 0.95, val * 1.05)
        others = [v for i, v in enumerate(values) if i != idx]
        if all(other < lower or other > upper for other in others):
          outliers.append(cust_names[idx])
      if outliers:
        outlier_result[pline] = outliers
    return outlier_result